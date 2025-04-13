import telebot
import datetime
import pandas as pd
import os
import glob
import re
from openpyxl import load_workbook
from win32com.client import Dispatch
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton


# 📌 Configuración de Telegram
TELEGRAM_BOT_TOKEN = "XXXXXXX"
bot = telebot.TeleBot(TELEGRAM_BOT_TOKEN)

data_file = "brine_data.csv"

# 📌 Ruta de imágenes
IMAGE_PATH = "brine_images/"

# 📌 Base de datos de productos
PRODUCTS = {
    "W500750P": {"Name": "FLATS FOR SHAVED BEEF","MAX INPUT %": 55, "TARGET RANGE %": "50-55*", "Nitrite": "No", "Allergen": "Soy\n"},
    "W500102P": {"Name": "BRISKET","MAX INPUT %": 78, "TARGET RANGE %": "60-65", "Nitrite": "No", "Allergen": "Soy\n"},
    "W10529P": {"Name": "SUBWAY BRISKET","MAX INPUT %": 70, "TARGET RANGE %": "63-68", "Nitrite": "No", "Allergen": "Soy\n"},
    "W10532P": {"Name": "FIREHOUSE BRISKET","MAX INPUT %": 70, "TARGET RANGE %": "60-65", "Nitrite": "NO", "Allergen": "None\n"},
    "W300009P1": {"Name": "PORK LOIN MM","MAX INPUT %": 84, "TARGET RANGE %": "57-62 Fresh/63-68 Defrost", "Nitrite": "Yes", "Allergen": "None\n"},
    "W300009NNP": {"Name": "NO NAME LOINS","MAX INPUT %": 76, "TARGET RANGE %": "57-62 Fresh/63-68 Defrost", "Nitrite": "Yes", "Allergen": "None\n"},
    "W300009NB": {"Name": "PORK LOIN NEUTRAL BRINE","MAX INPUT %": 81, "TARGET RANGE %": "57-62 Fresh/63-68 Defrost", "Nitrite": "Yes", "Allergen": "None\n"},
    "W300009P4": {"Name": "PORK LOIN CHOP","MAX INPUT %": 68, "TARGET RANGE %": "55-60", "Nitrite": "No", "Allergen": "None\n"},
    "W300009A": {"Name": "APPLEWOOD DOUBLE SMOKE","MAX INPUT %": 60, "TARGET RANGE %": "55-60", "Nitrite": "Yes", "Allergen": "None\n"},
    "W300009Q": {"Name": "DOUBLE SMOKE BACK BACON","MAX INPUT %": 60, "TARGET RANGE %": "55-60", "Nitrite": "Yes", "Allergen": "None\n"},
    "W33099P": {"Name": "FC DICED DBL SMOKE","MAX INPUT %": 60, "TARGET RANGE %": "55-60", "Nitrite": "Yes", "Allergen": "None\n"},
    "W38004": {"Name": "NO NAME","MAX INPUT %": 60, "TARGET RANGE %": "55-60", "Nitrite": "Yes", "Allergen": "None\n"},
    "W300009P2": {"Name": "NO NAME","MAX INPUT %": 22, "TARGET RANGE %": "20-22", "Nitrite": "No", "Allergen": "None\n"},
    "W300060PP": {"Name": "NO NAME","MAX INPUT %": 40, "TARGET RANGE %": "35-40", "Nitrite": "No", "Allergen": "None\n"},
    "W300064PP": {"Name": "PULLED PORK CUSHION","MAX INPUT %": 40, "TARGET RANGE %": "35-40", "Nitrite": "No", "Allergen": "None\n"},
    "W10406P": {"Name": "LIL JUANS CARNITAS","MAX INPUT %": 45, "TARGET RANGE %": "33-37", "Nitrite": "No", "Allergen": "None\n"},
    "W900510P": {"Name": "NO NAME","MAX INPUT %": 17, "TARGET RANGE %": "14-16", "Nitrite": "No", "Allergen": "None\n"},
    "W300100P": {"Name": "PORK BELLY BURNT ENDS","MAX INPUT %": 17, "TARGET RANGE %": "14-16", "Nitrite": "No", "Allergen": "None\n"}
}


# 📌 Datos adicionales para los productos específicos (NO SE MUESTRA EN PRODUCTS)
# BEEF FLATS
PRODUCT_DETAILS = {
    "W500750P": {
        "Brine Name": "BR00025",
        "Water": 70,
        "Bag Size": 8.7,
        "Total Batch": 78.7,
        "Percent": 0.60,
        "Max Bags per Tank": 10,
        "Image": "BR00025"
    },
# DOUBLE SMOMOKE
    "W300009Q": {
        "Brine Name": "BR00007",
        "Water": 146.44,
        "Bag Size": 19.02,
        "Bestate per Bag": 8.4,
        "Liquid Smoke per Bag": 0.124,
        "Total Batch": 173.98,
        "Percent": 0.65,
        "Max Bags per Tank": 5,
        "Image": "prueba.jpg"
    },
# APPLEWOOD
    "W300009A": {
        "Brine Name": "BR00014",
        "Water": 146.44,
        "Bag Size": 19.03,
        "Bestate per Bag": 8.4,
        "Liquid Smoke per Bag": 0.125,
        "Total Batch": 173.995,
        "Percent": 0.65,
        "Max Bags per Tank": 5,
        "Image": "prueba.jpg"
    },
# FIREHOUSE BRISKET
    "W10532P": {
        "Brine Name": "BR00020",
        "Water": 142.5,
        "Salt per Bag": 3.25,
        "Phosphate Ultra": 1.25,
        "Total Batch": 147,
        "Percent": 0.70,
        "Max Bags per Tank": 5,
        "Image": "prueba.jpg"
    },
# PORK LOIN MM
    "W300009P1": {
        "Brine Name": "BR00001",
        "Water": 154.57,
        "Bag Size": 12.62,
        "Phosphate Ultra": 0.9,
        "Total Batch": 168.09,
        "Percent": 0.90,
        "Max Bags per Tank": 5,
        "Image": "prueba.jpg"
    }
}

# 📌 Función para escapar caracteres especiales en MarkdownV2
def escape_markdown_v2(text):
    escape_chars = r'_*[]()~`>#+-=|{}.!'
    return ''.join(f'\\{char}' if char in escape_chars else char for char in text)

# 📌 Función para dividir mensajes largos
def split_message(message, max_length=4000):
    messages = []
    while len(message) > max_length:
        split_index = message[:max_length].rfind("\n")
        if split_index == -1:
            split_index = max_length
        messages.append(message[:split_index])
        message = message[split_index:].strip()
    messages.append(message)
    return messages




# 📌 Función para manejar mensajes en Telegram
@bot.message_handler(func=lambda message: True)
def handle_message(message):
    text = message.text.upper().strip()

    if text == "PRODUCTS-CODE":
        response = "📋 PRODUCTS LIST:\n\n"
        for code, details in PRODUCTS.items():
            response += f"🔹 {code}: Max Input {details['MAX INPUT %']}%, Target Range {details['TARGET RANGE %']}, Nitrite: {details['Nitrite']}, Allergen: {details['Allergen']}\n"
        response = escape_markdown_v2(response)
        for part in split_message(response):
            bot.send_message(message.chat.id, part, parse_mode="MarkdownV2")
        return
    
    if text.startswith("PRODUCT-"):
        code = text.replace("PRODUCT-", "").strip()
        if code in PRODUCTS:
            details = PRODUCTS[code]
            response = (f"📌INFORMATION ABOUT\n\n"
                        f"✔ Max Input %: {details['MAX INPUT %']}\n"
                        f"✔ Target Range %: {details['TARGET RANGE %']}\n"
                        f"✔ Nitrite: {details['Nitrite']}\n"
                        f"✔ Allergen: {details['Allergen']}")
        else:
            response = "❌ PRODUCT CODE NOT FOUND."
        response = escape_markdown_v2(response)
        bot.send_message(message.chat.id, response, parse_mode="MarkdownV2")
        return
    
    # 📌 Si el mensaje es "EMERGENCY", enviar el contacto del supervisor
    if text == "EMERGENCY":
        response = "🚨 CONTACT EMERGENCY:\n\n"
        response += "👨‍💼 *Barry Foss - Supervisor*\n📞 905 XXXXX"
        bot.send_message(message.chat.id, escape_markdown_v2(response), parse_mode="MarkdownV2")
        return

    # 📌 Si el mensaje es "TRANSFER", enviar los números de transferencia
    if text == "TRANSFER":
        response = f"📞 CODES TRANSFER:\n\n"
        response += "🔹 Pumping - 8687\n"
        response += "🔹 Brine - 8666\n"
        response += "🔹 Row Pack - 9999\n"
        response += "🔹 Cow Pack - 8888\n"
        response += "🔹 Defrost - 7777"
        bot.send_message(message.chat.id, escape_markdown_v2(response), parse_mode="MarkdownV2")
        return

    if text == "LOOK":
        response = (
            "PRODUCTS-CODE: Query product information.\n\n"
            "PRODUCT-W1XXXXX(code product): Query a specific product.\n\n"
            "TRANSFER: Codes to perform area transfers.\n\n"
            "EMERGENCY: Lous Kitchen emergency contacts.\n\n"
            "WXXXXX(code product)-1000(processed quantity): Returns all the necessary information to perform the product brine.\n\n"
            "LOOK: Query the consultation codes for the Telegram chat."
        )
    
        bot.send_message(message.chat.id, escape_markdown_v2(response), parse_mode="MarkdownV2")
        return


    if "-" in text:
        try:
            producto, cantidad = text.split("-")
            cantidad = float(cantidad)

            if producto not in PRODUCT_DETAILS:
                bot.send_message(message.chat.id, "❌ *Código de producto no encontrado.*", parse_mode="MarkdownV2")
                return

            response = f"🛠 *To prepare the correct brine for {producto}, we must use the following product(s):*"
            bot.send_message(message.chat.id, escape_markdown_v2(response), parse_mode="MarkdownV2")


            # 📌 Enviar todas las imágenes que coincidan con el Brine Name del producto
            brine_name = PRODUCT_DETAILS[producto]["Brine Name"]  # BR00025 o BR00007, etc.
            image_pattern = os.path.join(IMAGE_PATH, f"{brine_name}*")  # Busca todas las imágenes relacionadas con el Brine Name
            image_files = glob.glob(image_pattern)  # Obtiene la lista de archivos coincidentes

            if image_files:
                for img_path in image_files:
                    file_name = os.path.basename(img_path)  # Obtiene solo el nombre del archivo con la extensión
        
            # 📌 Extraer la parte después del guion y antes de la extensión
                    match = re.search(rf"{brine_name}-(.+?)\.", file_name)
                    if match:
                        extracted_name = match.group(1)  # Extrae el nombre entre "-" y "."
                    else:
                        extracted_name = file_name  # Si no hay coincidencia, usa el nombre completo

                    caption_text = f"🧂Name Product: `{extracted_name}`"

                    with open(img_path, "rb") as img:
                        bot.send_photo(message.chat.id, img, caption=escape_markdown_v2(caption_text), parse_mode="MarkdownV2")
            else:
                bot.send_message(message.chat.id, "⚠️ *Imágenes no encontradas.*")


            resultado = calcular_brine(producto, cantidad)

            df = pd.DataFrame([resultado])
            df.to_csv(data_file, mode='a', header=not pd.io.common.file_exists(data_file), index=False)

            response = "📊 *Resultado del Cálculo:*\n"
            response += "\n".join([f"✔ {k}: {v}" for k, v in resultado.items()])
            bot.send_message(message.chat.id, escape_markdown_v2(response), parse_mode="MarkdownV2")

            # 📌 Enviar el PDF si se generó
            if "PDF Generated" in resultado and resultado["PDF Generated"]:
                with open(resultado["PDF Generated"], "rb") as pdf_file:
                    bot.send_document(message.chat.id, pdf_file, caption="📄 *Brine Report Generated*")

        except Exception as e:
            bot.send_message(message.chat.id, escape_markdown_v2(f"❌ *Error en el formato.* Usa: CÓDIGO-CANTIDAD.\n🔹 *Ejemplo:* W500750P-6000\n\nError: {str(e)}"), parse_mode="MarkdownV2")
        return

# 📌 Función para calcular el brine
def calcular_brine(producto, producto_procesado):
    details = PRODUCT_DETAILS[producto]
    bag_size = details["Bag Size"]
    total_batch = details["Total Batch"]
    max_bags_per_tank = details["Max Bags per Tank"]
    percent = details["Percent"]

    bolsas_totales = round(round(producto_procesado * percent) + 250) // total_batch + 2
    tanques_requeridos = bolsas_totales // max_bags_per_tank
    bolsas_restantes = bolsas_totales % max_bags_per_tank

    if bolsas_restantes > 0:
        tanques_requeridos += 1

    distribucion_tanques = f"\n💡 *Should be used {tanques_requeridos} tank(s):*"
    if bolsas_restantes > 0:
        distribucion_tanques += f"\n- {tanques_requeridos - 1} tanks with {max_bags_per_tank} bags"
        distribucion_tanques += f"\n- 1 tank with {bolsas_restantes} bags"
    else:
        distribucion_tanques += f"\n- {tanques_requeridos} tanques con {max_bags_per_tank} bags"

    pasos = generar_pasos(producto)

    # 📌 Calcular valores específicos según el producto
    if producto == "W300009A":  # BRINE 14
        total_water = details["Water"] * bolsas_totales
        total_liquid_smoke = details["Liquid Smoke per Bag"] * bolsas_totales
        total_bestate = details["Bestate per Bag"] * bolsas_totales
        total_tmf_applewood = details["Bag Size"] * bolsas_totales
        total_batch = total_water + total_liquid_smoke + total_bestate + total_tmf_applewood

        return {
            "Date": datetime.datetime.now().strftime("%Y-%m-%d"),
            "Product": producto,
            "Brine Name": details["Brine Name"],
            "Processed Product (kg)": producto_procesado,
            "Total Water": total_water,
            "Total Liquid Smoke": total_liquid_smoke,
            "Total TMF Applewood": total_tmf_applewood,
            "Total Bestate": total_bestate,
            "Total Bags": bolsas_totales,
            "Bags Per Tank": max_bags_per_tank,
            "Total Batch": f"{total_batch}\n---------------------------------------------------",  
            "DISTRIBUTION TANKS": f"\n{distribucion_tanques}\n---------------------------------------------------",
            "STEPS": f"\n{pasos}"
        }

    elif producto == "W300009Q":  # BRINE 07
        total_water = details["Water"] * bolsas_totales
        total_liquid_smoke = details["Liquid Smoke per Bag"] * bolsas_totales
        total_bestate = details["Bestate per Bag"] * bolsas_totales
        total_tmf_double_smoke = details["Bag Size"] * bolsas_totales
        total_batch = total_water + total_liquid_smoke + total_bestate + total_tmf_double_smoke

        return {
            "Date": datetime.datetime.now().strftime("%Y-%m-%d"),
            "Product": producto,
            "Brine Name": details["Brine Name"],
            "Processed Product (kg)": producto_procesado,
            "Total Water": total_water,
            "Total Liquid Smoke": total_liquid_smoke,
            "Total TMF Double Smoke": total_tmf_double_smoke,
            "Total Bestate": total_bestate,
            "Total Bags": bolsas_totales,
            "Bags Per Tank": max_bags_per_tank,
            "Total Batch": f"{total_batch}\n---------------------------------------------------",  
            "DISTRIBUTION TANKS": f"\n{distribucion_tanques}\n---------------------------------------------------",
            "STEPS": f"\n{pasos}"
        }

    elif producto == "W500750P":  # BRINE 25
        total_water = details["Water"] * bolsas_totales
        total_tmf_rotisserie = details["Bag Size"] * bolsas_totales
        total_batch = total_tmf_rotisserie + total_water

        # 📌 Actualizar formato Excel y generar PDF
        ruta_pdf = actualizar_formato_brine(producto, producto_procesado, total_water, total_tmf_rotisserie,total_batch)


        return {
            "Date": datetime.datetime.now().strftime("%Y-%m-%d"),
            "Product": producto,
            "Brine Name": details["Brine Name"],
            "Processed Product (kg)": producto_procesado,
            "Total Water": total_water,
            "Total TMF Rotisserie": total_tmf_rotisserie,
            "Total Bags": bolsas_totales,
            "Bags Per Tank": max_bags_per_tank,
            "Total Batch": f"{total_batch}\n---------------------------------------------------",  
            "DISTRIBUTION TANKS": f"\n{distribucion_tanques}\n---------------------------------------------------",
            "STEPS": f"\n{pasos}",
            "PDF Generated": ruta_pdf  # 📌 Incluir el PDF generado en el resultado
        }

    elif producto == "W10532P":  # FIREHOUSE BRISKET
        total_salt = details["Salt per Bag"] * bolsas_totales
        total_phosphate_ultra = details["Phosphate Ultra"] * bolsas_totales
        total_water = details["Water"] * bolsas_totales
        total_batch = total_water + total_salt + total_phosphate_ultra

        return {
            "Date": datetime.datetime.now().strftime("%Y-%m-%d"),
            "Product": producto,
            "Brine Name": details["Brine Name"],
            "Processed Product (kg)": producto_procesado,
            "Total Salt": total_salt,
            "Total Phosphate Ultra": total_phosphate_ultra,
            "Total Water": total_water,
            "Total Batch": f"{total_batch}\n---------------------------------------------------",  
            "DISTRIBUTION TANKS": f"\n{distribucion_tanques}\n---------------------------------------------------",
            "STEPS": f"\n{pasos}"
        }

    elif producto == "W300009P1":  # BRINE 01 PORK LOIN MM
        total_water = details["Water"] * bolsas_totales
        total_tmf_rotisserie = details["Bag Size"] * bolsas_totales
        total_phosphate_ultra = details["Phosphate Ultra"] * bolsas_totales
        total_batch = total_tmf_rotisserie + total_water + total_phosphate_ultra

        return {
            "Date": datetime.datetime.now().strftime("%Y-%m-%d"),
            "Product": producto,
            "Brine Name": details["Brine Name"],
            "Processed Product (kg)": producto_procesado,
            "Total Water": total_water,
            "Total Phosphate Ultra": total_phosphate_ultra,
            "Total TMF Rotisserie": total_tmf_rotisserie,
            "Total Bags": bolsas_totales,
            "Bags Per Tank": max_bags_per_tank,
            "Total Batch": f"{total_batch}\n---------------------------------------------------",  
            "DISTRIBUTION TANKS": f"\n{distribucion_tanques}\n---------------------------------------------------",
            "STEPS": f"\n{pasos}"
        }

    # Si el producto no coincide con los especificados, retorna un resultado vacío
    return {
        "Date": datetime.datetime.now().strftime("%Y-%m-%d"),
        "Product": producto,
        "Messaje": "❌ Producto no encontrado en los cálculos específicos.",
        "\nDistribución de Tanques": distribucion_tanques,
        "\nSteps": pasos
    }

    #-------------------------------------------------------------------------------------------------------------------------



def actualizar_formato_brine(producto, producto_procesado, total_water, total_tmf_rotisserie, total_batch):
    # 📌 Verificar que sea el producto correcto
    if producto != "W500750P":
        print("❌ Este producto no usa el formato BR00025-Formato.xlsx")
        return
    
    # 📌 Definir la ruta del archivo
    fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d")
    ruta_archivo = os.path.join("Formatos-Output", "BR00025-Formato.xlsx")
    ruta_pdf = os.path.join("Formatos-Output", f"BR00025-Formato-{fecha_actual}.pdf")
    
    # 📌 Verificar si el archivo existe
    if not os.path.exists(ruta_archivo):
        print("❌ El archivo BR00025-Formato.xlsx no se encontró en la carpeta Formatos-Output.")
        return
    
    # 📌 Cargar el archivo de Excel
    wb = load_workbook(ruta_archivo)
    ws = wb.active  # Usar la hoja activa
    
    # 📌 Llenar las celdas con los valores correspondientes
    ws["E5"] = fecha_actual  # Fecha actual
    ws["C9"] = "BR00025"  # Código de brine
    ws["D9"] = "FLTAS FOR SHAVED BEEF"  # Nombre del producto
    ws["E9"] = producto_procesado  # Producto procesado
    ws["F14"] = total_water  # Total water
    ws["F15"] = total_tmf_rotisserie  # Total TMF Rotisserie
    ws["F17"] = total_batch
    
    # 📌 Guardar los cambios en Excel
    wb.save(ruta_archivo)
    wb.close()
    print("✅ Archivo BR00025-Formato.xlsx actualizado correctamente.")
    
    # 📌 Convertir el archivo de Excel a PDF
    try:
        excel = Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(os.path.abspath(ruta_archivo))
        ws = wb.ActiveSheet
        wb.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
        wb.Close(False)
        excel.Quit()
        print(f"✅ Archivo {ruta_pdf} generado correctamente.")
    except Exception as e:
        print(f"❌ Error al convertir Excel a PDF: {e}")
        return
    
    # 📌 Mostrar el PDF en el chat
    return ruta_pdf





    #-------------------------------------------------------------------------------------------------------------------------

# 📌 Función para generar los pasos del proceso
def generar_pasos(producto):
    if producto == "W300009Q":#DOUBLE SMOKE
        return ("\n🔄 *Brine7. Double Smoked Back Bacon*\n"
                "\n"
                "*Brine making procedure*\n"
                "\n"
                "1️⃣ *Inspect all tanks and equipment to ensure they are clean, free from defects, free from foreign material, and pose no threats to food safety.*\n"
                "2️⃣ *Add half of the total amount of required water. Check water temperature, and add ice if the temperature exceeds 4C. Add remaining water.*\n"
                "3️⃣ *Turn on the mixer. Slowly add 'Cooked pea meal Unit' blend for 5 minutes, until mixture appears clear.*\n"
                "4️⃣ *Slowly add Bestate (Lactate/ Diacetate), blend for an additional 5 minutes.*\n"
                "5️⃣ *Take a salometer reading twice and record. Take the brine temperature and glycol tank temperature, record.*\n"
                "6️⃣ *Continue blending brine until tank is empty.*\n")
                
            
    elif producto == "W500750P":#BEEF FLATS
        return ("\n🔄 *Brine25. Rotisserie (2024 Revised)*\n"
                "\n"
                "*Brine making procedure*\n"
                "\n"
                "1️⃣ *Inspect all tanks and equipment to ensure they are clean, free from defects, free from foreign material, and pose no threats to food safety.*\n"
                "2️⃣ *Add half of the total amount of required water. Check water temperature, and add ice if the temperature exceeds 4C. Add remaining water.*\n"
                "3️⃣ *Turn on the mixer. Slowly add 'TMF Rotisserie 2010', blend for 10minutes, until mixture appears clear*\n"
                "4️⃣ *Slowly add Bestate (Lactate/ Diacetate), blend for an additional 5 minutes.*\n"
                "5️⃣ *Take a salometer reading twice and record. Take the brine temperature and glycol tank temperature, record.*\n"
                "6️⃣ *Continue blending brine until tank is empty.*\n")
    
    elif producto == "W300009A":#APPLEWOOD
        return ("\n🔄 *Brine14. Applewood Smoked Back Bacon*\n"
                "\n"
                "*Brine making procedure*\n"
                "\n"
                "1️⃣ *Inspect all tanks and equipment to ensure they are clean, free from defects, free from foreign material, and pose no threats to food safety.*\n"
                "2️⃣ *Add half of the total amount of required water. Check water temperature, and add ice if the temperature exceeds 4C. Add remaining water.*\n"
                "3️⃣ *Turn on the mixer. Slowly add 'Cooked pea meal unit' blend 5 minutes, until mixture appears clear.*\n"
                "4️⃣ *Slowly add Bestate (Lactate/ Diacetate), blend for an additional 5 minutes.*\n"
                "5️⃣ *Take a salometer reading twice and record. Take the brine temperature and glycol tank temperature, record.*\n"
                "6️⃣ *Continue blending brine until tank is empty.*\n")
    
    elif producto == "W10532P":#FIREHOUSE BRISKET
        return ("\n🔄 *Brine20. Rotisserie (2024 Revised)*\n"
                "\n"
                "*Brine making procedure*\n"
                "\n"
                "1️⃣ *Inspect all tanks and equipment to ensure they are clean, free from defects, free from foreign material, and pose no threats to food safety.*\n"
                "2️⃣ *Add half of the total amount of required water. Check water temperature, and add ice if the temperature exceeds 4C. Add remaining water.*\n"
                "3️⃣ *Turn on the mixer. Slowly add Phosphate, blend for 5 minutes, until mixture appears clear.*\n"
                "4️⃣ *Slowly add salt, blend for 5 minutes.*\n"
                "5️⃣ *Take a salometer reading twice and record. Take brine temperature and glycol tank temperature and record.*\n"
                "6️⃣ *Continue blending brine until tank is empty.*\n"
                "7️⃣ *Ensure brine start and release times and all lot numbers for all materials used are recorded.*\n")
    
    elif producto == "W300009P1":#PORK LOIN MM
        return ("\n🔄 *Brine01. Fresh Peamel Brine (For P1)*\n"
                "\n"
                "*Brine making procedure*\n"
                "\n"
                "1️⃣ *Inspect all tanks and equipment to ensure they are clean, free from defects, free from foreign material, and pose no threats to food safety.*\n"
                "2️⃣ *Add half of the total amount of required water. Check water temperature, and add ice if the temperature exceeds 4C. Add remaining water.*\n"
                "3️⃣ *Turn on the mixer. Slowly add Phosphate, blend for 5 minutes, until mixture appears clear.*\n"
                "4️⃣ *Slowly add P1 brine until blend for 5 minutes, add all remaining ingredients. Blend for an additional 5 minutes*\n"
                "5️⃣ *Take a salometer reading twice and record. Take brine temperature and glycol tank temperature and record.*\n"
                "6️⃣ *Continue blending brine until tank is empty.*\n")

    return ""

    

# 📌 Iniciar el bot
bot.polling()
