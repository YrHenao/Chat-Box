import telebot
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
import datetime
import pandas as pd
import os
import glob
import re
import win32com.client
import base64
import json
import google.auth
import base64
from bs4 import BeautifulSoup
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
import html
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from win32com.client import Dispatch
from telebot.types import ReplyKeyboardMarkup, KeyboardButton

# 📌 Configurar el alcance de Gmail API (solo lectura)
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

# 📌 Configuración del bot de Telegram
TELEGRAM_BOT_TOKEN = "XXXXXXXXXX"
bot = telebot.TeleBot(TELEGRAM_BOT_TOKEN)

# Directorio donde se almacenan las imágenes para settins injector
IMAGE_FOLDER = "images-settings-injector"

# 📌 Ruta de imágenes
IMAGE_PATH = "brine_images/"

# Definir estados
STATE_SEARCH_PRODUCT = "search_product"
STATE_SEARCH_PRODUCTU = "search_product"
STATE_SEARCH_SETTINGS = "search_settings"

# Diccionario para almacenar el estado de cada usuario
user_states = {}

# 📌 Diccionario de productos
PRODUCTS = {
    "W500750P": {"Name": "FLATS FOR SHAVED BEEF","Max Input %": 55, "Target Range %": "50-55*", "Nitrite": "No", "Allergen": "Soy\n"},
    "W500102P": {"Name": "BRISKET","Max Input %": 78, "Target Range %": "60-65", "Nitrite": "No", "Allergen": "Soy\n"},
    "W10529P": {"Name": "SUBWAY BRISKET","Max Input %": 70, "Target Range %": "63-68", "Nitrite": "No", "Allergen": "Soy\n"},
    "W10532P": {"Name": "FIREHOUSE BRISKET","Max Input %": 70, "Target Range %": "60-65", "Nitrite": "NO", "Allergen": "None\n"},
    "W300009P1": {"Name": "PORK LOIN MM","Max Input %": 84, "Target Range %": "57-62 Fresh/63-68 Defrost", "Nitrite": "Yes", "Allergen": "None\n"},
    "W300009NNP": {"Name": "NO NAME LOINS","Max Input %": 76, "Target Range %": "57-62 Fresh/63-68 Defrost", "Nitrite": "Yes", "Allergen": "None\n"},
    "W300009NB": {"Name": "PORK LOIN NEUTRAL BRINE","Max Input %": 81, "Target Range %": "57-62 Fresh/63-68 Defrost", "Nitrite": "Yes", "Allergen": "None\n"},
    "W300009P4": {"Name": "PORK LOIN CHOP","Max Input %": 68, "Target Range %": "55-60", "Nitrite": "No", "Allergen": "None\n"},
    "W300009A": {"Name": "APPLEWOOD DOUBLE SMOKE","Max Input %": 60, "Target Range %": "55-60", "Nitrite": "Yes", "Allergen": "None\n"},
    "W300009Q": {"Name": "DOUBLE SMOKE BACK BACON","Max Input %": 60, "Target Range %": "55-60", "Nitrite": "Yes", "Allergen": "None\n"},
    "W33099P": {"Name": "FC DICED DBL SMOKE","Max Input %": 60, "Target Range %": "55-60", "Nitrite": "Yes", "Allergen": "None\n"},
    "W38004": {"Name": "NO NAME","Max Input %": 60, "Target Range %": "55-60", "Nitrite": "Yes", "Allergen": "None\n"},
    "W300009P2": {"Name": "NO NAME","Max Input %": 22, "Target Range %": "20-22", "Nitrite": "No", "Allergen": "None\n"},
    "W300060PP": {"Name": "NO NAME","Max Input %": 40, "Target Range %": "35-40", "Nitrite": "No", "Allergen": "None\n"},
    "W300064PP": {"Name": "PULLED PORK CUSHION","Max Input %": 40, "Target Range %": "35-40", "Nitrite": "No", "Allergen": "None\n"},
    "W10406P": {"Name": "LIL JUANS CARNITAS","Max Input %": 45, "Target Range %": "33-37", "Nitrite": "No", "Allergen": "None\n"},
    "W900510P": {"Name": "NO NAME","Max Input %": 17, "Target Range %": "14-16", "Nitrite": "No", "Allergen": "None\n"},
    "W300100P": {"Name": "PORK BELLY BURNT ENDS","Max Input %": 17, "Target Range %": "14-16", "Nitrite": "No", "Allergen": "None\n"}
}

PRODUCTSTU = {
    "W500750TU": {"Name": "FLATS FOR SHAVED BEEF","Max Input %": 55, "Target Range %": "50-55*", "Nitrite": "No", "Allergen": "Soy\n"},
    "W500102TU": {"Name": "BRISKET","Max Input %": 78, "Target Range %": "60-65", "Nitrite": "No", "Allergen": "Soy\n"},
    "W10529TU": {"Name": "SUBWAY BRISKET","Max Input %": 70, "Target Range %": "63-68", "Nitrite": "No", "Allergen": "Soy\n"},
    "W10532TU": {"Name": "FIREHOUSE BRISKET","Max Input %": 70, "Target Range %": "60-65", "Nitrite": "NO", "Allergen": "None\n"},
    "W300009P4": {"Name": "PORK LOIN CHOP","Max Input %": 68, "Target Range %": "55-60", "Nitrite": "No", "Allergen": "None\n"},
    "W10696TU": {"Name": "CARA BONE IN BEEF RIB","Max Input %": 60, "Target Range %": "55-60", "Nitrite": "Yes", "Allergen": "None\n"},
    "W300100P": {"Name": "PORK BELLY BURNT ENDS","Max Input %": 17, "Target Range %": "14-16", "Nitrite": "No", "Allergen": "None\n"}
}

TRANSFERCODE = {
    "8687": {"Room": "Pumping","Code": 8687},
    "8666": {"Room": "Brine","Code": 8666},
    "8": {"Room": "Tumbling","Code": 8},
    "8": {"Room": "Defrost","Code": 8},
    "8": {"Room": "Raw Pack","Code": 8},
    "8": {"Room": "Cook Pack","Code": 8},
    "8": {"Room": "Racking","Code": 8}
}

PRODUCTSBR = {
    "BR00001": {"Name": "Pork Loin MM","Max Input %": 80, "Allergen": "None"},
    "BR00007": {"Name": "Double Smoke Back Bacon","Max Input %": 65, "Allergen": "None"},
    "BR00005": {"Name": "No Name","Max Input %": 65, "Allergen": "None"},
    "BR00025": {"Name": "Beef Flats","Max Input %": 60, "Allergen": "Soy"},
    "BR00012": {"Name": "Firehouse Brisket","Max Input %": 70, "Allergen": "Soy"},
}

# 📌 Función para autenticar y conectar a Gmail
def authenticate_gmail():
    creds = None
    token_path = "token.json"
    credentials_path = "credentials.json"

    # Cargar credenciales guardadas
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    # Si no hay credenciales o han expirado, pedir autorización
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
            creds = flow.run_local_server(port=8080)

        # Guardar el token para futuros accesos
        with open(token_path, "w") as token:
            token.write(creds.to_json())

    return creds

# 📌 Función para dividir mensajes largos en partes manejables para Telegram
def split_message(text, max_length=4096):
    parts = []
    while len(text) > max_length:
        split_index = text[:max_length].rfind("\n")
        if split_index == -1:
            split_index = max_length
        parts.append(text[:split_index])
        text = text[split_index:].strip()
    parts.append(text)
    return parts

def escape_markdown_v2(text):
    """ Escapa caracteres especiales en MarkdownV2 para evitar errores en Telegram. """
    if not text:
        return ""

    # Caracteres especiales que deben escaparse en MarkdownV2
    escape_chars = r"_*[]()~`>#+-=|{}.!<>"
    
    # Escapar cada carácter especial con '\'
    escaped_text = re.sub(r"([_\*\[\]\(\)~`>#\+\-=|{}.!<>])", r"\\\1", text)

    return escaped_text






# 📌 Función mejorada para extraer TODO el historial del correo, incluyendo respuestas anteriores
def extract_email_body(payload):
    """ Extrae el cuerpo completo del correo, incluyendo toda la cadena de respuestas. """
    email_body = ""

    if "parts" in payload:
        for part in payload["parts"]:
            mime_type = part["mimeType"]
            body_data = part["body"].get("data", "")

            if body_data:
                decoded_body = base64.urlsafe_b64decode(body_data).decode("utf-8", errors="ignore")

                if mime_type == "text/plain":
                    email_body += decoded_body + "\n\n"
                elif mime_type == "text/html":
                    soup = BeautifulSoup(decoded_body, "html.parser")
                    email_body += soup.get_text() + "\n\n"

            # 📌 Revisar si hay partes anidadas dentro de la estructura del correo (respuestas anteriores)
            if "parts" in part:
                email_body += extract_email_body(part)  # Llamado recursivo para extraer más información

    else:
        # 📌 Si no tiene partes, se trata de un mensaje simple (texto plano)
        body_data = payload["body"].get("data", "")
        if body_data:
            email_body = base64.urlsafe_b64decode(body_data).decode("utf-8", errors="ignore")

    return email_body.strip()




# 📌 Función mejorada para obtener el correo más reciente con TODO el historial de respuestas
def search_gmail_latest_email(chat_id):
    try:
        creds = authenticate_gmail()
        service = build("gmail", "v1", credentials=creds)

        # 📌 Buscar el correo más reciente de "YrwinH@louskitchen.ca"
        query = 'from:YrwinH@louskitchen.ca'
        results = service.users().messages().list(userId="me", q=query, maxResults=1).execute()

        if "messages" in results:
            message_id = results["messages"][0]["id"]
            message = service.users().messages().get(userId="me", id=message_id, format="full").execute()

            # 📌 Obtener el asunto del correo y escaparlo
            subject = "Sin Asunto"
            for header in message["payload"]["headers"]:
                if header["name"] == "Subject":
                    subject = header["value"]
                    break

            # 📌 Extraer TODO el historial de correos
            email_body = extract_email_body(message["payload"])

            # 📌 Guardar el contenido en un archivo .txt con TODO el historial del correo
            file_path = "ultimo_correo.txt"  # Guardar en el mismo directorio del script
            with open(file_path, "w", encoding="utf-8") as file:
                file.write(f"📧 *Último correo recibido de YrwinH@louskitchen.ca*\n")
                file.write(f"Asunto: {subject}\n\n")
                file.write(email_body)  # 📌 Ahora contiene TODO el historial de la conversación

            # 📌 Enviar el archivo adjunto en Telegram
            with open(file_path, "rb") as doc:
                bot.send_document(chat_id, doc, caption="📧 *Último correo recibido de YrwinH@louskitchen.ca*")

            # 📌 Eliminar el archivo después de enviarlo (opcional)
            os.remove(file_path)

        else:
            bot.send_message(chat_id, "❌ No se encontró ningún correo reciente de YrwinH@louskitchen.ca.",
                             parse_mode="MarkdownV2", reply_markup=get_persistent_keyboard())

    except Exception as e:
        bot.send_message(chat_id, f"⚠️ *Error al buscar el correo en Gmail:* {escape_markdown_v2(str(e))}",
                         parse_mode="MarkdownV2", reply_markup=get_persistent_keyboard())



#  BOTONES PERSISTENTES 
def get_persistent_keyboard():
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
    home_button = KeyboardButton("🏠 Home")
    clear_button = KeyboardButton("🗑️ Limpiar Chat")
    markup.add(home_button, clear_button)
    return markup


#def get_persistent_keyboard():
#    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
#    home_button = KeyboardButton("🏠 Home")
#    clear_button = KeyboardButton("🗑️ Limpiar Chat")
#    markup.add(home_button, clear_button)
#    return markup

# 📌 Función para dividir mensajes largos
def split_message(message, max_length=4096):
    """
    Divide un mensaje en partes más pequeñas si excede max_length caracteres.
    """
    parts = []
    while len(message) > max_length:
        split_index = message.rfind('\n', 0, max_length)
        if split_index == -1:
            split_index = max_length
        parts.append(message[:split_index])
        message = message[split_index:]
    parts.append(message)
    return parts

# 📌 Función para mostrar el menú principal
def start_menu(message):
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("🚀 Pumping", callback_data="pumping"))
    markup.add(InlineKeyboardButton("🔄 Tumbling", callback_data="tumbling"))
    markup.add(InlineKeyboardButton("💧 Brine", callback_data="brine"))
    markup.add(InlineKeyboardButton("❄️ Defrost", callback_data="defrost"))
    markup.add(InlineKeyboardButton("🏭 Lous Kitchen", callback_data="lous_kitchen"))
    bot.send_message(
        message.chat.id,
        "Welcome To The Personal Bot", 
        parse_mode="Markdown", 
        reply_markup=markup
    )

     # Enviar teclado persistente con los botones inferiores
    bot.send_message(
        message.chat.id,
        "---Click To Option---", 
     #   "Usa estos botones para navegar rápidamente:",
        reply_markup=get_persistent_keyboard()
    )

@bot.message_handler(commands=['start', 'menu'])
def menu_handler(message):
    start_menu(message)

# 📌 Manejador de todas las callback_query
@bot.callback_query_handler(func=lambda call: True)
def callback_handler(call):
    if call.data == "pumping":
        show_pumping_menu(call.message.chat.id)
    elif call.data == "settings_injector":
        user_states[call.message.chat.id] = STATE_SEARCH_SETTINGS
        bot.send_message(call.message.chat.id, "🔢 *Please, put the product code to looking for picture:*", 
                         parse_mode="Markdown",reply_markup=get_persistent_keyboard())
    elif call.data == "tumbling":
        show_tumbling_menu(call.message.chat.id)
    elif call.data == "brine":
        show_brine_menu(call.message.chat.id)
    elif call.data == "defrost":
        bot.send_message(call.message.chat.id, "❄️ *Sección de Descongelación*\n\nInformación sobre el proceso de descongelación.", 
                         parse_mode="Markdown",reply_markup=get_persistent_keyboard())
    elif call.data == "lous_kitchen":
        show_lous_menu(call.message.chat.id)

    elif call.data == "products":#PUMPING
        show_products_menu(call.message.chat.id)
    elif call.data == "all_products":#PUMPING
        show_all_products(call.message.chat.id)
    elif call.data == "product_code":#PUMPING
        user_states[call.message.chat.id] = STATE_SEARCH_PRODUCT
        bot.send_message(call.message.chat.id, "🔍 Please, put here product code to looking for:", 
                         parse_mode="Markdown",reply_markup=get_persistent_keyboard())

    elif call.data == "productstu":#TUMBLING
        show_productstu_menu(call.message.chat.id)
    elif call.data == "all_productstu":#TUMBLING
        show_all_products(call.message.chat.id)
    elif call.data == "product_codetu":#TUMBLING
        user_states[call.message.chat.id] = STATE_SEARCH_PRODUCT
        bot.send_message(call.message.chat.id, "🔍 Please, put here code product to looking for:", 
                         parse_mode="Markdown",reply_markup=get_persistent_keyboard())
        
    elif call.data == "productsbr":#BRINE
        show_productsbr_menu(call.message.chat.id)
    elif call.data == "all_productsbr":#BRINE
        show_all_productsbr(call.message.chat.id)
    elif call.data == "product_codebr":#BRINE
        user_states[call.message.chat.id] = STATE_SEARCH_PRODUCT
        bot.send_message(call.message.chat.id, "🔍 Please, put here code product to looking for:", 
                         parse_mode="Markdown",reply_markup=get_persistent_keyboard())

    elif call.data == "about":
        bot.send_message(call.message.chat.id, "ℹ️ *Acerca de*\n\nInformación sobre el sistema Pumping.", 
                         parse_mode="Markdown",reply_markup=get_persistent_keyboard())
    elif call.data == "yield":
        bot.send_message(call.message.chat.id, "📊 *Yield*\n\nInformación sobre el rendimiento.", 
                         parse_mode="Markdown",reply_markup=get_persistent_keyboard())
    elif call.data == "injector":
        bot.send_message(call.message.chat.id, "🛠 *Programador Injector*\n\nHerramientas para programar el inyector.", 
                         parse_mode="Markdown",reply_markup=get_persistent_keyboard())
    elif call.data == "transfer_code":
        show_transfer_codes(call.message.chat.id)
    elif call.data == "gmail_email":
        bot.send_message(call.message.chat.id, "📧 Buscando el último correo en Gmail con 'production' en el asunto...",
                     parse_mode="Markdown", reply_markup=get_persistent_keyboard())
        search_gmail_latest_email(call.message.chat.id)
    elif call.data == "emergency":
        send_emergency_image(call.message.chat.id)
    elif call.data == "holidays":
        send_holidays_image(call.message.chat.id)
    elif call.data == "payroll":
        send_payroll_image(call.message.chat.id)




    else:
        bot.send_message(call.message.chat.id, "❓ *Opción no reconocida.*", parse_mode="Markdown",reply_markup=get_persistent_keyboard())

# 📌 FUNCION PARA ENVIAR IMAGEN DE EMERGENCY
def send_emergency_image(chat_id):
    """Envía la imagen de emergencia almacenada en la carpeta Formats-Company dentro del proyecto."""
    # 📌 Obtener la ruta del script actual
    script_directory = os.path.dirname(os.path.abspath(__file__))

    # 📌 Construir la ruta relativa a la imagen
    image_path = os.path.join(script_directory, "Formats-Company", "Call_In_Procedure_Clean.jpg")

    try:
        with open(image_path, "rb") as photo:
            bot.send_photo(chat_id, photo, caption="📌 *Call In Procedure*", parse_mode="Markdown")
    except Exception as e:
        bot.send_message(chat_id, f"⚠️ *Error al enviar la imagen:* {escape_markdown_v2(str(e))}", parse_mode="MarkdownV2")


# 📌 FUNCION PARA ENVIAR IMAGEN DE PAYROLL
def send_payroll_image(chat_id):
    """Envía la imagen de emergencia almacenada en la carpeta Formats-Company dentro del proyecto."""
    # 📌 Obtener la ruta del script actual
    script_directory = os.path.dirname(os.path.abspath(__file__))

    # 📌 Construir la ruta relativa a la imagen
    image_path = os.path.join(script_directory, "Formats-Company", "2025 Payroll Calendar.jpg")

    try:
        with open(image_path, "rb") as photo:
            bot.send_photo(chat_id, photo, caption="📌 *2025 Payroll Calendar*", parse_mode="Markdown")
    except Exception as e:
        bot.send_message(chat_id, f"⚠️ *Error al enviar la imagen:* {escape_markdown_v2(str(e))}", parse_mode="MarkdownV2")



# 📌 FUNCION PARA ENVIAR IMAGEN DE HOLIDAYS
def send_holidays_image(chat_id):
    """Envía la imagen de emergencia almacenada en la carpeta Formats-Company dentro del proyecto."""
    # 📌 Obtener la ruta del script actual
    script_directory = os.path.dirname(os.path.abspath(__file__))

    # 📌 Construir la ruta relativa a la imagen
    image_path = os.path.join(script_directory, "Formats-Company", "Statutory_Holidays_2025.jpg")

    try:
        with open(image_path, "rb") as photo:
            bot.send_photo(chat_id, photo, caption="📌 *Statutory Holidays*", parse_mode="Markdown")
    except Exception as e:
        bot.send_message(chat_id, f"⚠️ *Error al enviar la imagen:* {escape_markdown_v2(str(e))}", parse_mode="MarkdownV2")



# 📌 FUNCION PARA BUSCAR POR CODIGO DE PRODUCTO Y GENERE la imagen en settings injection
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == STATE_SEARCH_SETTINGS)
def handle_injector_code(message):
    code = message.text.strip().upper()
    
    # Buscar la imagen que comience con el código ingresado
    image_pattern = os.path.join(IMAGE_FOLDER, f"{code}-*.jpg")  # Busca archivos que empiecen con el código
    matching_images = glob.glob(image_pattern)

    if matching_images:
        with open(matching_images[0], "rb") as image:
            bot.send_photo(message.chat.id, image, caption=f"🖼️ Configuración para el código {code}",
                           reply_markup=get_persistent_keyboard())  # ✔ Muestra siempre los botones "Home" y "Limpiar Chat"
    else:
        bot.send_message(message.chat.id, "❌ No se encontró una imagen para ese código en la carpeta de configuraciones.", 
                         parse_mode="Markdown")

    user_states[message.chat.id] = None  # Restablecer el estado

# 📌 FUNCION PARA BUSCAR POR CODIGO DE PRODUCTO Y GENERE UN SOLO RESULTADO     PUMPINGGGG
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == STATE_SEARCH_PRODUCT)
def handle_product_code(message):
    code = message.text.strip().upper()
    if code in PRODUCTS:
        details = PRODUCTS[code]
        response = (f"📌 *About Of Product:*\n\n"
                    f"*Códe:* {escape_markdown_v2(code)}\n"
                    f"*Name:* {escape_markdown_v2(details['Name'])}\n"
                    f"*Max Input %:* {details['Max Input %']}\n"
                    f"*Target Range %:* {escape_markdown_v2(details['Target Range %'])}\n"
                    f"*Nitrite:* {escape_markdown_v2(details['Nitrite'])}\n"
                    f"*Allérgen:* {escape_markdown_v2(details['Allergen'])}")
    else:
        response = "❌ *Código de producto no encontrado.*"
    bot.send_message(message.chat.id, response, parse_mode="MarkdownV2", reply_markup=get_persistent_keyboard())
    user_states[message.chat.id] = None  # Restablecer el estado del usuario

# 📌 FUNCION PARA BUSCAR POR CODIGO DE PRODUCTO Y GENERE UN SOLO RESULTADO        TUMBLING
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == STATE_SEARCH_PRODUCTU)
def handle_product_code(message):
    code = message.text.strip().upper()
    if code in PRODUCTSTU:
        details = PRODUCTSTU[code]
        response = (f"📌 *About Of Product:*\n\n"
                    f"*Códe:* {escape_markdown_v2(code)}\n"
                    f"*Name:* {escape_markdown_v2(details['Name'])}\n"
                    f"*Max Input %:* {details['Max Input %']}\n"
                    f"*Target Range %:* {escape_markdown_v2(details['Target Range %'])}\n"
                    f"*Nitrite:* {escape_markdown_v2(details['Nitrite'])}\n"
                    f"*Allérgen:* {escape_markdown_v2(details['Allergen'])}")
    else:
        response = "❌ *Código de producto no encontrado.*"
    bot.send_message(message.chat.id, response, parse_mode="MarkdownV2", reply_markup=get_persistent_keyboard())
    user_states[message.chat.id] = None  # Restablecer el estado del usuario





# FUNCION PARA BRINE CALCULATE ------------------------------------------------------------------------------------------------------

PRODUCT_DETAILS = {
    "W500750P": {
        "Brine Name": "BR00025",
        "Water": 70,
        "Bag Size": 8.7,
        "Total Batch": 78.7,
        "Percent": 0.60,
        "Max Bags per Tank": 10,
        "Image": "BR00025"
    },
# DOUBLE SMOMOKE
    "W300009Q": {
        "Brine Name": "BR00007",
        "Water": 146.44,
        "Bag Size": 19.02,
        "Bestate per Bag": 8.4,
        "Liquid Smoke per Bag": 0.124,
        "Total Batch": 173.98,
        "Percent": 0.65,
        "Max Bags per Tank": 5,
        "Image": "prueba.jpg"
    },
# APPLEWOOD
    "W300009A": {
        "Brine Name": "BR00014",
        "Water": 146.44,
        "Bag Size": 19.03,
        "Bestate per Bag": 8.4,
        "Liquid Smoke per Bag": 0.125,
        "Total Batch": 173.995,
        "Percent": 0.65,
        "Max Bags per Tank": 5,
        "Image": "prueba.jpg"
    },
# FIREHOUSE BRISKET
    "W10532P": {
        "Brine Name": "BR00020",
        "Water": 142.5,
        "Salt per Bag": 3.25,
        "Phosphate Ultra": 1.25,
        "Total Batch": 147,
        "Percent": 0.70,
        "Max Bags per Tank": 5,
        "Image": "prueba.jpg"
    },
# PORK LOIN MM
    "W300009P1": {
        "Brine Name": "BR00001",
        "Water": 154.57,
        "Bag Size": 12.62,
        "Phosphate Ultra": 0.9,
        "Total Batch": 168.09,
        "Percent": 0.90,
        "Max Bags per Tank": 5,
        "Image": "prueba.jpg"
    }
}

@bot.callback_query_handler(func=lambda call: call.data == "brine_calculate")
def prompt_brine_calculation(call):
    bot.send_message(call.message.chat.id, "Por favor, ingresa el código de producto y la cantidad en el siguiente formato:\n\n`CÓDIGO-CANTIDAD`\n\n*Ejemplo:* `W500750P-6000`", parse_mode="Markdown")
    bot.register_next_step_handler(call.message, process_brine_input)

    # 📌 Guardamos el estado del usuario para saber que debe ingresar el código y la cantidad
    user_states[call.message.chat.id] = "waiting_for_brine_input"


# 📌 Manejador para capturar la respuesta del usuario después de ingresar el código y cantidad
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == "waiting_for_brine_input")
def process_brine_input(message):
    text = message.text.strip()
    if "-" in text:
        try:
            producto, cantidad = text.split("-")
            cantidad = float(cantidad)

            if producto not in PRODUCT_DETAILS:
                bot.send_message(message.chat.id, "❌ *Código de producto no encontrado.*", parse_mode="Markdown")
                return

            response = f"🛠 *Para preparar la salmuera correcta para {producto}, se deben utilizar los siguientes productos:*"
            bot.send_message(message.chat.id, response, parse_mode="Markdown")

            # 📌 Enviar imágenes relacionadas con el producto
            brine_name = PRODUCT_DETAILS[producto]["Brine Name"]
            image_pattern = os.path.join(IMAGE_PATH, f"{brine_name}*")
            image_files = glob.glob(image_pattern)

            if image_files:
                for img_path in image_files:
                    with open(img_path, "rb") as img:
                        bot.send_photo(message.chat.id, img)
            else:
                bot.send_message(message.chat.id, "⚠️ *Imágenes no encontradas.*", parse_mode="Markdown")

            # 📌 Realizar cálculos de la salmuera
            resultado = calcular_brine(producto, cantidad)

            # 📌 Enviar resultados al usuario
            response = "📊 *Resultado del Cálculo:*\n"
            for k, v in resultado.items():
                response += f"✔ {k}: {v}\n"
            bot.send_message(message.chat.id, response, parse_mode="Markdown")

            # 📌 Enviar el PDF generado, si existe
            if "PDF Generated" in resultado and resultado["PDF Generated"]:
                with open(resultado["PDF Generated"], "rb") as pdf_file:
                    bot.send_document(message.chat.id, pdf_file, caption="📄 *Informe de Salmuera Generado*", parse_mode="Markdown")

        except Exception as e:
            bot.send_message(message.chat.id, f"❌ *Error en el formato.* Usa: CÓDIGO-CANTIDAD.\n🔹 *Ejemplo:* W500750P-6000\n\nError: {str(e)}", parse_mode="Markdown")
    else:
        bot.send_message(message.chat.id, "❌ *Formato incorrecto.* Asegúrate de usar el formato: CÓDIGO-CANTIDAD.\n🔹 *Ejemplo:* W500750P-6000", parse_mode="Markdown")

def calcular_brine(producto, producto_procesado):
    details = PRODUCT_DETAILS[producto]
    bag_size = details["Bag Size"]
    total_batch = details["Total Batch"]
    max_bags_per_tank = details["Max Bags per Tank"]
    percent = details["Percent"]

    bolsas_totales = round((producto_procesado * percent) / total_batch)
    tanques_requeridos = bolsas_totales // max_bags_per_tank
    bolsas_restantes = bolsas_totales % max_bags_per_tank

    if bolsas_restantes > 0:
        tanques_requeridos += 1

    distribucion_tanques = f"\n💡 *Se deben usar {tanques_requeridos} tanque(s):*"
    if bolsas_restantes > 0:
        distribucion_tanques += f"\n- {tanques_requeridos - 1} tanques con {max_bags_per_tank} bolsas"
        distribucion_tanques += f"\n- 1 tanque con {bolsas_restantes} bolsas"
    else:
        distribucion_tanques += f"\n- {tanques_requeridos} tanques con {max_bags_per_tank} bolsas"

    pasos = generar_pasos(producto)

    # 📌 Calcular valores específicos según el producto
    if producto == "W300009A":  # BRINE 14
        total_water = details["Water"] * bolsas_totales
        total_liquid_smoke = details["Liquid Smoke per Bag"] * bolsas_totales
        total_bestate = details["Bestate per Bag"] * bolsas_totales
        total_tmf_applewood = details["Bag Size"] * bolsas_totales
        total_batch = total_water + total_liquid_smoke + total_bestate + total_tmf_applewood

        return {
            "Fecha": datetime.datetime.now().strftime("%Y-%m-%d"),
            "Producto": producto,
            "Nombre de Salmuera": details["Brine Name"],
            "Producto Procesado (kg)": producto_procesado,
            "Agua Total": total_water,
            "Total de Liquid Smoke": total_liquid_smoke,
            "Total de TMF Applewood": total_tmf_applewood,
            "Total de Bestate": total_bestate,
            "Total de Bolsas": bolsas_totales,
            "Bolsas por Tanque": max_bags_per_tank,
            "Total Batch": f"{total_batch}\n---------------------------------------------------",  
            "DISTRIBUTION TANKS": f"\n{distribucion_tanques}\n---------------------------------------------------",
            "STEPS": f"\n{pasos}"
        }
    
    elif producto == "W300009Q":  # BRINE 07
        total_water = details["Water"] * bolsas_totales
        total_liquid_smoke = details["Liquid Smoke per Bag"] * bolsas_totales
        total_bestate = details["Bestate per Bag"] * bolsas_totales
        total_tmf_double_smoke = details["Bag Size"] * bolsas_totales
        total_batch = total_water + total_liquid_smoke + total_bestate + total_tmf_double_smoke

        return {
            "Date": datetime.datetime.now().strftime("%Y-%m-%d"),
            "Product": producto,
            "Brine Name": details["Brine Name"],
            "Processed Product (kg)": producto_procesado,
            "Total Water": total_water,
            "Total Liquid Smoke": total_liquid_smoke,
            "Total TMF Double Smoke": total_tmf_double_smoke,
            "Total Bestate": total_bestate,
            "Total Bags": bolsas_totales,
            "Bags Per Tank": max_bags_per_tank,
            "Total Batch": f"{total_batch}\n---------------------------------------------------",  
            "DISTRIBUTION TANKS": f"\n{distribucion_tanques}\n---------------------------------------------------",
            "STEPS": f"\n{pasos}"
        }

    elif producto == "W500750P":  # BRINE 25
        total_water = details["Water"] * bolsas_totales
        total_tmf_rotisserie = details["Bag Size"] * bolsas_totales
        total_batch = total_tmf_rotisserie + total_water

        # 📌 Actualizar formato Excel y generar PDF
        ruta_pdf = actualizar_formato_brine(producto, producto_procesado, total_water, total_tmf_rotisserie,total_batch)


        return {
            "Date": datetime.datetime.now().strftime("%Y-%m-%d"),
            "Product": producto,
            "Brine Name": details["Brine Name"],
            "Processed Product (kg)": producto_procesado,
            "Total Water": total_water,
            "Total TMF Rotisserie": total_tmf_rotisserie,
            "Total Bags": bolsas_totales,
            "Bags Per Tank": max_bags_per_tank,
            "Total Batch": f"{total_batch}\n---------------------------------------------------",  
            "DISTRIBUTION TANKS": f"\n{distribucion_tanques}\n---------------------------------------------------",
            "STEPS": f"\n{pasos}",
            "PDF Generated": ruta_pdf  # 📌 Incluir el PDF generado en el resultado
        }

    elif producto == "W10532P":  # FIREHOUSE BRISKET
        total_salt = details["Salt per Bag"] * bolsas_totales
        total_phosphate_ultra = details["Phosphate Ultra"] * bolsas_totales
        total_water = details["Water"] * bolsas_totales
        total_batch = total_water + total_salt + total_phosphate_ultra

        return {
            "Date": datetime.datetime.now().strftime("%Y-%m-%d"),
            "Product": producto,
            "Brine Name": details["Brine Name"],
            "Processed Product (kg)": producto_procesado,
            "Total Salt": total_salt,
            "Total Phosphate Ultra": total_phosphate_ultra,
            "Total Water": total_water,
            "Total Batch": f"{total_batch}\n---------------------------------------------------",  
            "DISTRIBUTION TANKS": f"\n{distribucion_tanques}\n---------------------------------------------------",
            "STEPS": f"\n{pasos}"
        }

    elif producto == "W300009P1":  # BRINE 01 PORK LOIN MM
        total_water = details["Water"] * bolsas_totales
        total_tmf_rotisserie = details["Bag Size"] * bolsas_totales
        total_phosphate_ultra = details["Phosphate Ultra"] * bolsas_totales
        total_batch = total_tmf_rotisserie + total_water + total_phosphate_ultra

        return {
            "Date": datetime.datetime.now().strftime("%Y-%m-%d"),
            "Product": producto,
            "Brine Name": details["Brine Name"],
            "Processed Product (kg)": producto_procesado,
            "Total Water": total_water,
            "Total Phosphate Ultra": total_phosphate_ultra,
            "Total TMF Rotisserie": total_tmf_rotisserie,
            "Total Bags": bolsas_totales,
            "Bags Per Tank": max_bags_per_tank,
            "Total Batch": f"{total_batch}\n---------------------------------------------------",  
            "DISTRIBUTION TANKS": f"\n{distribucion_tanques}\n---------------------------------------------------",
            "STEPS": f"\n{pasos}"
        }

    # Si el producto no coincide con los especificados, retorna un resultado vacío
    return {
        "Date": datetime.datetime.now().strftime("%Y-%m-%d"),
        "Product": producto,
        "Messaje": "❌ Producto no encontrado en los cálculos específicos.",
        "\nDistribución de Tanques": distribucion_tanques,
        "\nSteps": pasos
    }    


def actualizar_formato_brine(producto, producto_procesado, total_water, total_tmf_rotisserie, total_batch):
    # 📌 Verificar que sea el producto correcto
    if producto != "W500750P":
        print("❌ Este producto no usa el formato BR00025-Formato.xlsx")
        return
    
    # 📌 Definir la ruta del archivo
    fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d")
    ruta_archivo = os.path.join("Formatos-Output", "BR00025-Formato.xlsx")
    ruta_pdf = os.path.join("Formatos-Output", f"BR00025-Formato-{fecha_actual}.pdf")
    
    # 📌 Verificar si el archivo existe
    if not os.path.exists(ruta_archivo):
        print("❌ El archivo BR00025-Formato.xlsx no se encontró en la carpeta Formatos-Output.")
        return
    
    # 📌 Cargar el archivo de Excel
    wb = load_workbook(ruta_archivo)
    ws = wb.active  # Usar la hoja activa
    
    # 📌 Llenar las celdas con los valores correspondientes
    ws["E5"] = fecha_actual  # Fecha actual
    ws["C9"] = "BR00025"  # Código de brine
    ws["D9"] = "FLTAS FOR SHAVED BEEF"  # Nombre del producto
    ws["E9"] = producto_procesado  # Producto procesado
    ws["F14"] = total_water  # Total water
    ws["F15"] = total_tmf_rotisserie  # Total TMF Rotisserie
    ws["F17"] = total_batch
    
    # 📌 Guardar los cambios en Excel
    wb.save(ruta_archivo)
    wb.close()
    print("✅ Archivo BR00025-Formato.xlsx actualizado correctamente.")
    
    # 📌 Convertir el archivo de Excel a PDF
    try:
        excel = Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(os.path.abspath(ruta_archivo))
        ws = wb.ActiveSheet
        wb.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
        wb.Close(False)
        excel.Quit()
        print(f"✅ Archivo {ruta_pdf} generado correctamente.")
    except Exception as e:
        print(f"❌ Error al convertir Excel a PDF: {e}")
        return
    
    # 📌 Mostrar el PDF en el chat
    return ruta_pdf



# 📌 Función para generar los pasos del proceso
def generar_pasos(producto):
    if producto == "W300009Q":#DOUBLE SMOKE
        return ("\n🔄 *Brine7. Double Smoked Back Bacon*\n"
                "\n"
                "*Brine making procedure*\n"
                "\n"
                "1️⃣ *Inspect all tanks and equipment to ensure they are clean, free from defects, free from foreign material, and pose no threats to food safety.*\n"
                "2️⃣ *Add half of the total amount of required water. Check water temperature, and add ice if the temperature exceeds 4C. Add remaining water.*\n"
                "3️⃣ *Turn on the mixer. Slowly add 'Cooked pea meal Unit' blend for 5 minutes, until mixture appears clear.*\n"
                "4️⃣ *Slowly add Bestate (Lactate/ Diacetate), blend for an additional 5 minutes.*\n"
                "5️⃣ *Take a salometer reading twice and record. Take the brine temperature and glycol tank temperature, record.*\n"
                "6️⃣ *Continue blending brine until tank is empty.*\n")
                
            
    elif producto == "W500750P":#BEEF FLATS
        return ("\n🔄 *Brine25. Rotisserie (2024 Revised)*\n"
                "\n"
                "*Brine making procedure*\n"
                "\n"
                "1️⃣ *Inspect all tanks and equipment to ensure they are clean, free from defects, free from foreign material, and pose no threats to food safety.*\n"
                "2️⃣ *Add half of the total amount of required water. Check water temperature, and add ice if the temperature exceeds 4C. Add remaining water.*\n"
                "3️⃣ *Turn on the mixer. Slowly add 'TMF Rotisserie 2010', blend for 10minutes, until mixture appears clear*\n"
                "4️⃣ *Slowly add Bestate (Lactate/ Diacetate), blend for an additional 5 minutes.*\n"
                "5️⃣ *Take a salometer reading twice and record. Take the brine temperature and glycol tank temperature, record.*\n"
                "6️⃣ *Continue blending brine until tank is empty.*\n")
    
    elif producto == "W300009A":#APPLEWOOD
        return ("\n🔄 *Brine14. Applewood Smoked Back Bacon*\n"
                "\n"
                "*Brine making procedure*\n"
                "\n"
                "1️⃣ *Inspect all tanks and equipment to ensure they are clean, free from defects, free from foreign material, and pose no threats to food safety.*\n"
                "2️⃣ *Add half of the total amount of required water. Check water temperature, and add ice if the temperature exceeds 4C. Add remaining water.*\n"
                "3️⃣ *Turn on the mixer. Slowly add 'Cooked pea meal unit' blend 5 minutes, until mixture appears clear.*\n"
                "4️⃣ *Slowly add Bestate (Lactate/ Diacetate), blend for an additional 5 minutes.*\n"
                "5️⃣ *Take a salometer reading twice and record. Take the brine temperature and glycol tank temperature, record.*\n"
                "6️⃣ *Continue blending brine until tank is empty.*\n")
    
    elif producto == "W10532P":#FIREHOUSE BRISKET
        return ("\n🔄 *Brine20. Rotisserie (2024 Revised)*\n"
                "\n"
                "*Brine making procedure*\n"
                "\n"
                "1️⃣ *Inspect all tanks and equipment to ensure they are clean, free from defects, free from foreign material, and pose no threats to food safety.*\n"
                "2️⃣ *Add half of the total amount of required water. Check water temperature, and add ice if the temperature exceeds 4C. Add remaining water.*\n"
                "3️⃣ *Turn on the mixer. Slowly add Phosphate, blend for 5 minutes, until mixture appears clear.*\n"
                "4️⃣ *Slowly add salt, blend for 5 minutes.*\n"
                "5️⃣ *Take a salometer reading twice and record. Take brine temperature and glycol tank temperature and record.*\n"
                "6️⃣ *Continue blending brine until tank is empty.*\n"
                "7️⃣ *Ensure brine start and release times and all lot numbers for all materials used are recorded.*\n")
    
    elif producto == "W300009P1":#PORK LOIN MM
        return ("\n🔄 *Brine01. Fresh Peamel Brine (For P1)*\n"
                "\n"
                "*Brine making procedure*\n"
                "\n"
                "1️⃣ *Inspect all tanks and equipment to ensure they are clean, free from defects, free from foreign material, and pose no threats to food safety.*\n"
                "2️⃣ *Add half of the total amount of required water. Check water temperature, and add ice if the temperature exceeds 4C. Add remaining water.*\n"
                "3️⃣ *Turn on the mixer. Slowly add Phosphate, blend for 5 minutes, until mixture appears clear.*\n"
                "4️⃣ *Slowly add P1 brine until blend for 5 minutes, add all remaining ingredients. Blend for an additional 5 minutes*\n"
                "5️⃣ *Take a salometer reading twice and record. Take brine temperature and glycol tank temperature and record.*\n"
                "6️⃣ *Continue blending brine until tank is empty.*\n")

    return ""




#------------------------------------------------------------------------------------------------------------------------------------


# MANEJADORES BOTONES PERSISTENTES
@bot.message_handler(func=lambda message: message.text == "🏠 Home")
def return_to_main_menu(message):
    start_menu(message)

@bot.message_handler(func=lambda message: message.text == "🗑️ Limpiar Chat")
def clear_chat(message):
    bot.send_message(
        message.chat.id,
        "🗑️ Para limpiar el chat, elimina manualmente los mensajes.",
        parse_mode="Markdown",
        reply_markup=get_persistent_keyboard()
    )

# COLOCAR AQUI LOS SUBMENUS CORRESPONDIENTES A CADA ROOM
# 📌 Submenú de Pumping
def show_pumping_menu(chat_id):
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("ℹ️ About Of Me", url="https://www.google.com/search?q=BRINE+PUMPING&sca_esv=e74a10221791e7e3&rlz=1C1CHBF_esCO1075CO1075&sxsrf=AHTn8zokMRAVfTRJ8qzUSWzXlYIhSQH83w%3A1740448127895&ei=fyG9Z7GRNu79ptQPze7fuAM&ved=0ahUKEwjxzKbD2t2LAxXuvokEHU33FzcQ4dUDCBA&uact=5&oq=BRINE+PUMPING&gs_lp=Egxnd3Mtd2l6LXNlcnAiDUJSSU5FIFBVTVBJTkcyBRAAGIAEMgYQABgWGB4yBhAAGBYYHjILEAAYgAQYhgMYigUyCxAAGIAEGIYDGIoFMgUQABjvBTIFEAAY7wVI0kRQAFiLI3AAeAGQAQCYAXqgAeMJqgEDOC41uAEDyAEA-AEBmAINoAKVCsICChAjGIAEGCcYigXCAgQQIxgnwgILEAAYgAQYkQIYigXCAgsQLhiABBiRAhiKBcICDhAAGIAEGLEDGIMBGIoFwgIOEC4YgAQYsQMY0QMYxwHCAhEQLhiABBixAxjRAxiDARjHAcICBRAuGIAEwgIKEAAYgAQYQxiKBcICChAuGIAEGEMYigXCAg4QLhiABBiRAhjJAxiKBcICCxAAGIAEGJIDGIoFwgINEC4YgAQYsQMYQxiKBcICDRAAGIAEGLEDGEMYigXCAgoQABiABBgUGIcCwgIIEAAYgAQYogSYAwCSBwM4LjWgB92dAQ&sclient=gws-wiz-serp"))
    markup.add(InlineKeyboardButton("📦 Products", callback_data="products"))
    markup.add(InlineKeyboardButton("🛠 Settings Injector", callback_data="settings_injector"))
    markup.add(InlineKeyboardButton("📋 Transfer Code", callback_data="transfer_code"))
    bot.send_message(chat_id, " ▐▐▐  PUMPING ▐▐▐ ", parse_mode="Markdown", reply_markup=markup)

# 📌 Submenú de Tumbling
def show_tumbling_menu(chat_id):
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("ℹ️ About Of Me", url="https://www.google.com/search?q=TUMBLING+FOOD+INFORMACION&rlz=1C1CHBF_esCO1075CO1075&oq=TUMBLING+FOOD+INFORMACION&gs_lcrp=EgZjaHJvbWUyBggAEEUYOdIBCTExNDQ3ajBqN6gCCLACAQ&sourceid=chrome&ie=UTF-8"))
    markup.add(InlineKeyboardButton("📦 Products", callback_data="productstu"))
    markup.add(InlineKeyboardButton("📊 Tumblers", callback_data="yield"))
    markup.add(InlineKeyboardButton("📋 Transfer Code", callback_data="transfer_code"))
    bot.send_message(chat_id, " ▐▐▐  TUMBLING ▐▐▐ ", parse_mode="Markdown", reply_markup=markup)

# 📌 Submenú de Brine
def show_brine_menu(chat_id):
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("ℹ️ About Of Me", url="https://www.google.com/search?q=BRINE+FOOD&sca_esv=e74a10221791e7e3&rlz=1C1CHBF_esCO1075CO1075&sxsrf=AHTn8zri9rM937v9Vf_KP4o5r8vWdzbQAQ%3A1740448158050&ei=niG9Z5njAtz9ptQP4encsA0&ved=0ahUKEwiZqtfR2t2LAxXcvokEHeE0F9YQ4dUDCBA&uact=5&oq=BRINE+FOOD&gs_lp=Egxnd3Mtd2l6LXNlcnAiCkJSSU5FIEZPT0QyBRAuGIAEMgUQABiABDIFEAAYgAQyBRAAGIAEMgYQABgWGB4yBhAAGBYYHjIGEAAYFhgeMgYQABgWGB4yBhAAGBYYHjIGEAAYFhgeMhQQLhiABBiXBRjcBBjeBBjfBNgBAUiqGlC2EFjGFHABeAGQAQCYAYQBoAGRA6oBAzMuMbgBA8gBAPgBAZgCBaACrAPCAgoQABiwAxjWBBhHwgILEAAYgAQYkQIYigXCAgoQLhiABBhDGIoFwgIKEAAYgAQYFBiHApgDAIgGAZAGBLoGBggBEAEYFJIHAzQuMaAHqD0&sclient=gws-wiz-serp"))
    markup.add(InlineKeyboardButton("📦 Products", callback_data="productsbr"))
    markup.add(InlineKeyboardButton("📊 Brine Calculate", callback_data="brine_calculate"))
    markup.add(InlineKeyboardButton("📋 Transfer Code", callback_data="transfer_code"))
    bot.send_message(chat_id, " ▐▐▐  BRINE ▐▐▐ ", parse_mode="Markdown", reply_markup=markup)

# 📌 Submenú de LOUS KITCHEN
def show_lous_menu(chat_id):
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("🏭 About of Me", url="https://louskitchen.ca/story/"))
    markup.add(InlineKeyboardButton("Calendar", callback_data="payroll"))
    markup.add(InlineKeyboardButton("📆 Holidays", callback_data="holidays"))
    markup.add(InlineKeyboardButton("🚨 Emergency", callback_data="emergency"))
    markup.add(InlineKeyboardButton("📜 Etic Code", callback_data="yield"))
    markup.add(InlineKeyboardButton("📧 Outlook", callback_data="gmail_email"))
    bot.send_message(chat_id, " ▐▐▐ 🏭 LOUS KITCHEN ▐▐▐ ", parse_mode="Markdown", reply_markup=markup)

# 📌 LISTA DE TRANSFER ESTO ES PARA PUMPING, TUMBLING, BRINE----------------------------------------------------------------------

def show_transfer_codes(chat_id):
    """Muestra la lista de códigos de transferencia disponibles."""
    message = "📋 *Lista de Transfer Codes:*\n\n"
    for code, details in TRANSFERCODE.items():
        message += f"✅ *Code:* {code} - *Room:* {details['Room']}\n"

    bot.send_message(chat_id, message, parse_mode="Markdown")


# 📌 Menú de Productos ESTO ES PARA BRINE---------------------------------------------------------------------------------------
def show_productsbr_menu(chat_id):
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("📜 All Products", callback_data="all_productsbr"))
    markup.add(InlineKeyboardButton("🔍 Search Code", callback_data="product_codebr"))
    bot.send_message(chat_id, "📦 ↓↓↓ Click Home To Return ↓↓↓", parse_mode="Markdown", reply_markup=markup)

def show_all_productsbr(chat_id):
    """Muestra la lista de productos disponibles en Brine."""
    message = "📦 *Lista de Productos Brine:*\n\n"
    for code, details in PRODUCTSBR.items():
        message += f"✅ *Código:* {code}\n"
        message += f"🔹 *Nombre:* {details['Name']}\n"
        message += f"🔹 *Max Input %:* {details['Max Input %']}%\n"
        message += f"🔹 *Alérgeno:* {details['Allergen']}\n"
        message += "\n"  # Espaciado entre productos

    bot.send_message(chat_id, message, parse_mode="Markdown")

# 📌 Menú de Productos ESTO ES PARA PUMPING---------------------------------------------------------------------------------------
def show_products_menu(chat_id):
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("📜 All Products", callback_data="all_products"))
    markup.add(InlineKeyboardButton("🔍 Search Code", callback_data="product_code"))
    bot.send_message(chat_id, "📦 ↓↓↓ Click Home To Return ↓↓↓", parse_mode="Markdown", reply_markup=markup)

# Función para mostrar todos los productos PUMPING
def show_all_products(chat_id):
    # Construir el mensaje con la lista de productos
    message = "📦 *List Of Access Products:*\n\n"
    for code, details in PRODUCTS.items():
        message += f"*Códe:* {escape_markdown_v2(code)}\n"
        message += f"*Name:* {escape_markdown_v2(details['Name'])}\n"
        message += f"*Max Input %:* {details['Max Input %']}\n"
        message += f"*Target Ranget %:* {escape_markdown_v2(details['Target Range %'])}\n"
        message += f"*Nitrite:* {escape_markdown_v2(details['Nitrite'])}\n"
        message += f"*Allérgen:* {escape_markdown_v2(details['Allergen'])}\n"
        message += "\n"  # Línea en blanco para separar productos
    # Dividir el mensaje si excede el límite de longitud
    messages = split_message(message)
    # Enviar cada parte del mensaje al usuario
    for part in messages:
        bot.send_message(chat_id, part, parse_mode="MarkdownV2")

# 📌 Menú de Productos ESTO ES PARA TUMBLING---------------------------------------------------------------------------------------
def show_productstu_menu(chat_id):
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("📜 All Products", callback_data="all_productstu"))
    markup.add(InlineKeyboardButton("🔍 Search Code", callback_data="product_codetu"))
    bot.send_message(chat_id, "📦 ↓↓↓ Click Home To Return ↓↓↓", parse_mode="Markdown", reply_markup=markup)

# Función para mostrar todos los productos TUMBLING
def show_all_products(chat_id):
    # Construir el mensaje con la lista de productos
    message = "📦 *List Of Access Products:*\n\n"
    for code, details in PRODUCTSTU.items():
        message += f"*Códe:* {escape_markdown_v2(code)}\n"
        message += f"*Name:* {escape_markdown_v2(details['Name'])}\n"
        message += f"*Max Input %:* {details['Max Input %']}\n"
        message += f"*Target Ranget %:* {escape_markdown_v2(details['Target Range %'])}\n"
        message += f"*Nitrite:* {escape_markdown_v2(details['Nitrite'])}\n"
        message += f"*Allérgen:* {escape_markdown_v2(details['Allergen'])}\n"
        message += "\n"  # Línea en blanco para separar productos
    # Dividir el mensaje si excede el límite de longitud
    messages = split_message(message)
    # Enviar cada parte del mensaje al usuario
    for part in messages:
        bot.send_message(chat_id, part, parse_mode="MarkdownV2")

# Función para dividir mensajes largos en partes más pequeñas
def split_message(message, max_length=4096):
    return [message[i:i + max_length] for i in range(0, len(message), max_length)]







# 📌 Iniciar el bot
bot.polling()
